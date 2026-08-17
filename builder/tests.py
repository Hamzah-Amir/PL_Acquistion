"""Backend tests.

The heavy end-to-end test runs the real sample archive through the whole
pipeline and asserts the workflow §4 reconciliations hold.  It is skipped
automatically when the sample files are not present.
"""
from __future__ import annotations

import os
import shutil
import tempfile
import unittest
from pathlib import Path

from django.conf import settings
from django.test import TestCase, override_settings

from .engine import assemble, model as plmodel
from .models import Job
from .parsers import summary_pdf, transactions
from . import services

SAMPLES = Path(settings.BASE_DIR) / "sample_files"
SAMPLE_ARCHIVE = SAMPLES / "New zip.zip"
TEMPLATE = Path(settings.PL_TEMPLATE_PATH)


class VatSplitTests(TestCase):
    """Workflow §2 — the two halves must always re-sum to the inclusive figure."""

    def test_split_is_lossless(self):
        for amount in (-1234.56, -0.01, -10160.77, 0.0, 60.42, -7.77):
            net = plmodel.ex_vat(amount)
            vat = plmodel.vat_part(amount)
            self.assertAlmostEqual(net + vat, round(amount, 2), places=2)

    def test_vat_is_twenty_percent_of_net(self):
        amount = -1200.00
        self.assertAlmostEqual(plmodel.ex_vat(amount), -1000.00, places=2)
        self.assertAlmostEqual(plmodel.vat_part(amount), -200.00, places=2)


class GrossSalesTests(TestCase):
    """Sales tax collected sits alongside product sales; the VAT formulas need
    the VAT-inclusive figure."""

    def test_gross_sales_adds_the_tax_collected(self):
        row = plmodel.MonthRow(key="2025-08", label="Aug-25")
        row.product_sales = 10000.0
        row.sales_tax = 2000.0
        self.assertAlmostEqual(row.net_sales, 10000.0, places=2)
        self.assertAlmostEqual(row.gross_sales, 12000.0, places=2)

    def test_output_vat_uses_gross_not_net(self):
        """Dividing by 6 only extracts VAT correctly from a gross figure."""
        gross_based = plmodel._vat_line("20", 12000.0, 0.0)
        self.assertAlmostEqual(gross_based, -2000.0, places=2)
        # 12,000 gross = 10,000 net + 2,000 VAT, so the VAT really is 2,000.
        self.assertAlmostEqual(-gross_based, 12000.0 - 12000.0 / 1.2, places=2)

    def test_net_sales_still_ties_to_income_when_tax_is_present(self):
        row = plmodel.MonthRow(key="2025-08", label="Aug-25")
        row.product_sales = 10000.0
        row.sales_tax = 2000.0
        row.income_total = 10000.0
        self.assertAlmostEqual(row.income_difference, 0.0, places=2)


class VatSchemeTests(TestCase):
    """The three schemes the user specified."""

    def test_zero_rate_has_no_output_vat(self):
        self.assertEqual(plmodel._vat_line("0", 100000.0, -5000.0), 0.0)

    def test_flat_rate_is_sales_over_107_5_times_7_5(self):
        self.assertAlmostEqual(
            plmodel._vat_line("7.5", 217528.12, -35000.0), -15176.38, places=2
        )
        # The input VAT argument must not affect the flat-rate charge at all.
        self.assertEqual(
            plmodel._vat_line("7.5", 217528.12, -35000.0),
            plmodel._vat_line("7.5", 217528.12, 0.0),
        )

    def test_standard_is_sales_over_six_less_input_vat(self):
        # 60,000 / 6 = 10,000 output, less 4,000 reclaimable = 6,000 payable.
        self.assertAlmostEqual(
            plmodel._vat_line("20", 60000.0, -4000.0), -6000.0, places=2
        )

    def test_standard_goes_negative_when_input_exceeds_output(self):
        """A repayment due from HMRC, not a floored zero."""
        line = plmodel._vat_line("20", 60000.0, -12000.0)
        self.assertAlmostEqual(line, 2000.0, places=2)
        self.assertGreater(line, 0)

    def test_input_vat_is_never_double_counted(self):
        """On 0% and 7.5% the input VAT is a memo; profit must equal net sales
        plus the cost lines and the output VAT only."""
        for scheme in ("0", "7.5", "20"):
            row = plmodel.MonthRow(key="2025-08", label="Aug-25")
            row.product_sales = 10000.0
            row.referral_ex, row.referral_vat = -1000.0, -200.0
            row.fba_ex, row.fba_vat = -2000.0, -400.0
            row.ppc_ex, row.ppc_vat = -500.0, -100.0
            row.storage = -600.0
            row.other_fees = -120.0
            row.cogs = -3000.0
            row.input_vat = round(
                row.referral_vat + row.fba_vat + row.ppc_vat
                + plmodel.vat_of_inclusive(row.storage)
                + plmodel.vat_of_inclusive(row.other_fees), 2
            )
            row.output_vat = plmodel._vat_line(scheme, row.net_sales, row.input_vat)
            expected = round(
                row.net_sales + row.total_fees + row.total_ppc
                + row.cogs + row.output_vat + row.opex, 2
            )
            self.assertAlmostEqual(row.net_profit, expected, places=2, msg=scheme)

    def test_cogs_vat_included_versus_excluded(self):
        """Included -> COGS/6 is the import VAT. Excluded -> rate x 20% is added
        on top, which comes to the same VAT on the grossed-up figure."""
        parse = {
            "month_keys": ["2025-08"],
            "months": {"2025-08": {"label": "Aug-25"}},
            "transactions": {"2025-08": {"net_units": 100}},
            "business_reports": [],
            "invoices": [],
            "skus": [
                {
                    "sku": "X", "title": "", "family": "F",
                    "units_by_month": {"2025-08": 100},
                    "sales_by_month": {"2025-08": 1000.0},
                }
            ],
        }
        base = {"family_rates": {"F": 10.0}, "sku_families": {"X": "F"}, "vat_scheme": "0"}

        included = plmodel.build_model(
            parse, {**base, "cogs_vat_paid": True, "cogs_vat_included": True}
        )
        self.assertAlmostEqual(included.months[0].cogs, -1000.00, places=2)
        self.assertAlmostEqual(included.months[0].cogs_vat, -166.67, places=2)

        excluded = plmodel.build_model(
            parse, {**base, "cogs_vat_paid": True, "cogs_vat_included": False}
        )
        self.assertAlmostEqual(excluded.months[0].cogs, -1200.00, places=2)
        # rate x 20% = 10 x 0.2 x 100 units = 200
        self.assertAlmostEqual(excluded.months[0].cogs_vat, -200.00, places=2)

    def test_no_vat_paid_on_goods_uses_the_rate_untouched(self):
        parse = {
            "month_keys": ["2025-08"],
            "months": {"2025-08": {"label": "Aug-25"}},
            "transactions": {"2025-08": {"net_units": 100}},
            "business_reports": [],
            "invoices": [],
            "skus": [
                {
                    "sku": "X", "title": "", "family": "F",
                    "units_by_month": {"2025-08": 100},
                    "sales_by_month": {"2025-08": 1000.0},
                }
            ],
        }
        model = plmodel.build_model(
            parse,
            {
                "family_rates": {"F": 10.0}, "sku_families": {"X": "F"},
                "vat_scheme": "20", "cogs_vat_paid": False,
            },
        )
        month = model.months[0]
        self.assertAlmostEqual(month.cogs, -1000.00, places=2)
        self.assertEqual(month.cogs_vat, 0.0)
        # Nothing from COGS may leak into the reclaim.
        self.assertAlmostEqual(month.input_vat, 0.0, places=2)

    def test_included_flag_is_ignored_when_no_vat_was_paid(self):
        parse = {
            "month_keys": ["2025-08"],
            "months": {"2025-08": {"label": "Aug-25"}},
            "transactions": {"2025-08": {"net_units": 100}},
            "business_reports": [],
            "invoices": [],
            "skus": [
                {
                    "sku": "X", "title": "", "family": "F",
                    "units_by_month": {"2025-08": 100},
                    "sales_by_month": {"2025-08": 1000.0},
                }
            ],
        }
        base = {"family_rates": {"F": 10.0}, "sku_families": {"X": "F"},
                "vat_scheme": "20", "cogs_vat_paid": False}
        for included in (True, False):
            model = plmodel.build_model(parse, {**base, "cogs_vat_included": included})
            self.assertAlmostEqual(model.months[0].cogs, -1000.00, places=2)

    def test_cogs_vat_feeds_input_vat(self):
        row = plmodel.MonthRow(key="2025-08", label="Aug-25")
        row.cogs = -1200.0
        row.cogs_vat = plmodel.vat_of_inclusive(row.cogs)
        self.assertAlmostEqual(row.cogs_vat, -200.0, places=2)

    def test_legacy_boolean_flag_still_resolves(self):
        self.assertEqual(plmodel.normalise_scheme({"vat_registered": True}), "20")
        self.assertEqual(plmodel.normalise_scheme({"vat_registered": False}), "0")
        self.assertEqual(plmodel.normalise_scheme({"vat_scheme": "7.5"}), "7.5")


class FilenameMonthTests(TestCase):
    """A single-month export must be assigned to its own month, not bucketed by
    posted date — Amazon windows the export on local time but stamps rows UTC."""

    def test_single_month_export(self):
        self.assertEqual(
            transactions.filename_month("2025Aug1-2025Aug31CustomTransaction.csv"),
            "2025-08",
        )

    def test_multi_month_export_falls_back(self):
        self.assertIsNone(
            transactions.filename_month("2022Feb1-2026Aug12CustomTransaction.csv")
        )

    def test_unrecognised_name(self):
        self.assertIsNone(transactions.filename_month("transactions.csv"))


class FamilyProposalTests(TestCase):
    def test_promotional_variants_fold_into_the_parent_sku(self):
        mapping = assemble._families_from_skus(
            [
                ("B-001", 12000),
                ("amzn.gr.B-001-ZtOYD3HIjWvLmvQubyjqrnD-VG", 12),
                ("anglebrackets001", 15000),
                ("amzn.gr.anglebrackets001-G_sefJRNSvkx-VG", 10),
            ]
        )
        self.assertEqual(
            mapping["amzn.gr.B-001-ZtOYD3HIjWvLmvQubyjqrnD-VG"], mapping["B-001"]
        )
        self.assertEqual(
            mapping["amzn.gr.anglebrackets001-G_sefJRNSvkx-VG"],
            mapping["anglebrackets001"],
        )

    def test_colour_variants_share_a_family(self):
        mapping = assemble._families_from_skus(
            [("15k-pink-loom", 900), ("15k-loom-purple", 800), ("15k loom blue", 700)]
        )
        self.assertEqual(len(set(mapping.values())), 1)


class ReportMonthGuessTests(TestCase):
    def test_reports_are_ranked_onto_months_by_units(self):
        reports = [
            {"source": "a.csv", "units_ordered": 3000},
            {"source": "b.csv", "units_ordered": 1000},
            {"source": "c.csv", "units_ordered": 2000},
        ]
        txn = {
            "2025-07": {"order_units": 1010, "net_units": 1000},
            "2025-08": {"order_units": 2020, "net_units": 2000},
            "2025-09": {"order_units": 3030, "net_units": 3000},
        }
        mapping = assemble.suggest_report_months(reports, txn, list(txn))
        self.assertEqual(mapping["b.csv"], "2025-07")
        self.assertEqual(mapping["c.csv"], "2025-08")
        self.assertEqual(mapping["a.csv"], "2025-09")


class UploadNameTests(TestCase):
    def test_path_traversal_is_stripped(self):
        self.assertEqual(services.safe_filename("../../etc/passwd"), "passwd")
        self.assertEqual(services.safe_filename(r"C:\evil\x.zip"), "x.zip")
        self.assertNotIn("/", services.safe_filename("a/b/c.csv"))


class PageTests(TestCase):
    """Every template must render, and the upload form must reject an empty post."""

    def test_upload_page_renders(self):
        response = self.client.get("/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Upload the seller")
        self.assertContains(response, "dropzone")

    def test_upload_requires_a_file(self):
        response = self.client.post("/", {}, follow=True)
        self.assertContains(response, "at least one file")
        self.assertEqual(Job.objects.count(), 0)

    def test_progress_endpoint_returns_json(self):
        job = Job.objects.create()
        response = self.client.get(f"/job/{job.pk}/progress/")
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertIn("progress", payload)
        self.assertIs(payload["done"], False)

    def test_wizard_steps_redirect_until_the_parse_is_ready(self):
        job = Job.objects.create()
        for step in ("review", "costs", "assumptions", "result"):
            response = self.client.get(f"/job/{job.pk}/{step}/")
            self.assertRedirects(
                response, f"/job/{job.pk}/processing/", fetch_redirect_response=False
            )

    def test_download_404s_before_the_build(self):
        job = Job.objects.create()
        self.assertEqual(self.client.get(f"/job/{job.pk}/download/").status_code, 404)

    def test_discard_removes_the_job(self):
        job = Job.objects.create()
        response = self.client.post(f"/job/{job.pk}/discard/", follow=True)
        self.assertEqual(response.status_code, 200)
        self.assertFalse(Job.objects.filter(pk=job.pk).exists())


@unittest.skipUnless(
    SAMPLE_ARCHIVE.exists() and TEMPLATE.exists(),
    "sample_files/New zip.zip and the template are required",
)
class EndToEndTests(TestCase):
    """The golden rule: every month ties to Amazon's own totals."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.tmp = tempfile.mkdtemp(prefix="pl-test-")
        with override_settings(PL_WORKSPACE_ROOT=Path(cls.tmp)):
            cls.job = Job.objects.create()
            cls.job.ensure_workspace()
            shutil.copy(SAMPLE_ARCHIVE, cls.job.uploads_dir / "sources.zip")
            services._process(str(cls.job.pk))
            cls.job.refresh_from_db()

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.tmp, ignore_errors=True)
        super().tearDownClass()

    def setUp(self):
        self.job.refresh_from_db()

    def test_parse_succeeded(self):
        self.assertEqual(self.job.status, Job.Status.READY, self.job.error)
        self.assertEqual(len(self.job.parse["month_keys"]), 13)
        self.assertEqual(self.job.parse["errors"], [])

    def test_every_file_kind_was_recognised(self):
        counts = self.job.parse["counts"]
        self.assertEqual(counts["summary"], 13)
        self.assertEqual(counts["transaction"], 13)
        self.assertEqual(counts["business"], 13)
        self.assertGreaterEqual(counts["invoice"], 100)

    def test_income_and_expenses_reconcile_for_every_month(self):
        for key, month in self.job.parse["months"].items():
            self.assertAlmostEqual(month["income_difference"], 0.0, places=2, msg=key)
            self.assertAlmostEqual(month["expenses_difference"], 0.0, places=2, msg=key)

    def test_model_passes_the_blocking_checks(self):
        model = services.build_model(self.job)
        failures = [c.name for c in model.checks if c.severity == "error" and not c.passed]
        self.assertEqual(failures, [], f"blocking checks failed: {failures}")

    def test_cogs_reconciles_to_the_flat_rate(self):
        """Σ(units × rate) ÷ Σ units must return the rate itself — this catches
        unit-sign bugs between the COGS build and the P&L unit count."""
        choices = services.default_choices(self.job.parse, {})
        choices["family_rates"] = {name: 2.5 for name in choices["family_rates"]}
        self.job.choices = choices
        self.job.save(update_fields=["choices"])
        model = services.build_model(self.job)
        self.assertAlmostEqual(model.blended_landed_cost, 2.5, places=2)

    def _with_scheme(self, scheme):
        choices = services.default_choices(self.job.parse, self.job.choices)
        choices["vat_scheme"] = scheme
        self.job.choices = choices
        self.job.save(update_fields=["choices"])
        return services.build_model(self.job)

    def test_zero_scheme_has_no_output_vat_but_reports_input_vat(self):
        model = self._with_scheme("0")
        self.assertEqual(model.ttm_sum("output_vat"), 0.0)
        self.assertLess(model.ttm_sum("input_vat"), 0.0)

    def test_flat_scheme_matches_the_stated_formula(self):
        model = self._with_scheme("7.5")
        expected = -round(model.ttm_sum("gross_sales") / 107.5 * 7.5, 2)
        self.assertAlmostEqual(model.ttm_sum("output_vat"), expected, delta=1.0)

    def test_sales_tax_is_read_and_grosses_up_sales(self):
        model = self._with_scheme("20")
        self.assertGreater(model.ttm_sum("sales_tax"), 0.0)
        self.assertAlmostEqual(
            model.ttm_sum("gross_sales"),
            model.ttm_sum("net_sales") + model.ttm_sum("sales_tax"),
            delta=1.0,
        )
        # UK VAT at 20% means the tax should be roughly a sixth of gross.
        share = model.ttm_sum("sales_tax") / model.ttm_sum("gross_sales")
        self.assertGreater(share, 0.10)
        self.assertLess(share, 0.20)

    def test_standard_scheme_nets_off_input_vat(self):
        model = self._with_scheme("20")
        gross = -round(model.ttm_sum("gross_sales") / 6, 2)
        self.assertAlmostEqual(
            model.ttm_sum("output_vat"),
            round(gross - model.ttm_sum("input_vat"), 2),
            delta=1.0,
        )
        # Netting off the reclaim must cost less than the gross output VAT.
        self.assertGreater(model.ttm_sum("output_vat"), gross)

    def test_input_vat_covers_all_fees_and_ppc(self):
        """Input VAT = referral + FBA + PPC VAT plus storage and other fees / 6."""
        model = self._with_scheme("20")
        for month in model.months:
            expected = round(
                month.referral_vat + month.fba_vat + month.ppc_vat
                + round(month.storage / 6, 2) + round(month.other_fees / 6, 2), 2
            )
            self.assertAlmostEqual(month.input_vat, expected, places=2, msg=month.label)

    def test_changing_scheme_only_moves_the_vat_line(self):
        """Sales, fees, PPC and COGS must be identical across schemes."""
        base = self._with_scheme("0")
        for scheme in ("7.5", "20"):
            other = self._with_scheme(scheme)
            for attribute in ("net_sales", "gross_sales", "total_fees", "total_ppc",
                              "cogs", "input_vat"):
                self.assertAlmostEqual(
                    base.ttm_sum(attribute), other.ttm_sum(attribute),
                    places=2, msg=f"{attribute} moved on scheme {scheme}",
                )

    def test_ppc_invoices_are_deduplicated_and_split_at_twenty_percent(self):
        invoices = self.job.parse["invoices"]
        numbers = [i["number"] for i in invoices]
        self.assertEqual(len(numbers), len(set(numbers)))
        self.assertTrue(all(i["vat_rate_ok"] for i in invoices))

    def test_workbook_builds_and_opens(self):
        import openpyxl

        with override_settings(PL_WORKSPACE_ROOT=Path(self.tmp)):
            path = services.build_output(self.job)
        self.assertTrue(os.path.exists(path))

        wb = openpyxl.load_workbook(path)
        for sheet in (
            "Status & Missing", "Inputs_Monthly", "Inputs_Units", "Inputs_Traffic",
            "P&L", "SKU Concentration", "COGS Build", "PPC Invoices",
            "Fee Reconciliation", "Traffic & Conversion", "Summary",
        ):
            self.assertIn(sheet, wb.sheetnames)

        # The 13 months must land in Inputs_Monthly rows 2-14.
        monthly = wb["Inputs_Monthly"]
        self.assertEqual(
            sum(1 for row in range(2, 15) if monthly.cell(row, 1).value), 13
        )
        # Column AC carries the sales tax Amazon reports separately.
        self.assertEqual(monthly.cell(1, 29).value, "Sales tax collected")

    def test_wizard_walks_through_every_step(self):
        """Drive the real views the way the browser does."""
        with override_settings(PL_WORKSPACE_ROOT=Path(self.tmp)):
            job_id = self.job.pk

            review = self.client.get(f"/job/{job_id}/review/")
            self.assertEqual(review.status_code, 200)
            self.assertContains(review, "Review what was found")

            months = self.job.parse["month_keys"]
            payload = {"window": months}
            for report, key in self.job.parse["business_report_map"].items():
                payload[f"report__{report}"] = key
            self.assertRedirects(
                self.client.post(f"/job/{job_id}/review/", payload),
                f"/job/{job_id}/costs/", fetch_redirect_response=False,
            )

            costs = self.client.get(f"/job/{job_id}/costs/")
            self.assertEqual(costs.status_code, 200)
            self.assertContains(costs, "preview-data")
            self.job.refresh_from_db()
            families = sorted({
                self.job.choices["sku_families"][s["sku"]]
                for s in self.job.parse["skus"]
            })
            payload = {f"rate__{name}": "2.50" for name in families}
            payload["cogs_vat"] = "included"
            for sku in self.job.parse["skus"]:
                payload[f"family__{sku['sku']}"] = \
                    self.job.choices["sku_families"][sku["sku"]]
            self.assertRedirects(
                self.client.post(f"/job/{job_id}/costs/", payload),
                f"/job/{job_id}/assumptions/", fetch_redirect_response=False,
            )

            page = self.client.get(f"/job/{job_id}/assumptions/")
            self.assertEqual(page.status_code, 200)
            self.assertContains(page, "VAT scheme")
            self.assertRedirects(
                self.client.post(f"/job/{job_id}/assumptions/", {
                    "vat_scheme": "20", "opex_monthly": "600",
                    "asking_price": "60000", "business_name": "Test Ltd",
                }),
                f"/job/{job_id}/result/", fetch_redirect_response=False,
            )

            result = self.client.get(f"/job/{job_id}/result/")
            self.assertEqual(result.status_code, 200)
            self.assertContains(result, "chart-data")
            self.assertContains(result, "Download workbook")

            download = self.client.get(f"/job/{job_id}/download/")
            self.assertEqual(download.status_code, 200)
            self.assertIn("attachment", download["Content-Disposition"])
            self.assertGreater(
                len(b"".join(download.streaming_content)), 20000
            )

    def test_key_metrics_rows_are_present(self):
        import openpyxl

        with override_settings(PL_WORKSPACE_ROOT=Path(self.tmp)):
            path = services.build_output(self.job)
        pl = openpyxl.load_workbook(path)["P&L"]
        labels = {
            (pl.cell(row, 1).value or "").strip(): row for row in range(40, 54)
        }
        for expected in (
            "Referral / Selling fee %", "FBA fulfilment fee %",
            "Total Amazon fee % (subtotal)", "PPC / TACoS %", "COGS %",
            "Output VAT %", "Net margin %", "TOTAL (must be 100%)",
        ):
            self.assertIn(expected, labels, f"missing key metric row: {expected}")
        # No label may start with "=" — Excel would evaluate it as a formula.
        for row in range(1, 70):
            value = pl.cell(row, 1).value
            if isinstance(value, str):
                self.assertFalse(
                    value.lstrip().startswith("="),
                    f"P&L A{row} label looks like a formula: {value!r}",
                )
