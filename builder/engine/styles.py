"""Shared look-and-feel for the generated workbook.

One place for the palette, fonts, number formats and the small helpers that
build titles, header rows, section bands and bordered tables — so every tab
reads the same way.
"""
from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --- palette ---------------------------------------------------------------
NAVY = "1F3864"        # table header fill
BAND = "D9E2F3"        # section band fill
TOTAL = "E8EDF5"       # total / TTM fill
ZEBRA = "F7F9FC"       # alternating body rows
YELLOW = "FFF2CC"      # editable assumption
GREY = "8A8A8A"        # note text
RULE = "B4C6E7"        # borders

# --- fonts (workflow §2 colour key) ----------------------------------------
BASE = "Calibri"
INPUT = Font(name=BASE, size=10, color="0000CC")          # blue  = hard input
FORMULA = Font(name=BASE, size=10, color="000000")        # black = formula
LINK = Font(name=BASE, size=10, color="007A3D")           # green = cross-sheet
BODY = Font(name=BASE, size=10)
BOLD = Font(name=BASE, size=10, bold=True)
BOLD_LINK = Font(name=BASE, size=10, bold=True, color="007A3D")
HEADER = Font(name=BASE, size=10, bold=True, color="FFFFFF")
TITLE = Font(name=BASE, size=14, bold=True, color=NAVY)
SUBTITLE = Font(name=BASE, size=9, italic=True, color=GREY)
SECTION = Font(name=BASE, size=10, bold=True, color=NAVY)
NOTE = Font(name=BASE, size=9, italic=True, color=GREY)
TOTAL_FONT = Font(name=BASE, size=10, bold=True)

# --- fills -----------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
BAND_FILL = PatternFill("solid", fgColor=BAND)
TOTAL_FILL = PatternFill("solid", fgColor=TOTAL)
ZEBRA_FILL = PatternFill("solid", fgColor=ZEBRA)
YELLOW_FILL = PatternFill("solid", fgColor=YELLOW)
NO_FILL = PatternFill(fill_type=None)

# --- borders ---------------------------------------------------------------
_THIN = Side(style="thin", color=RULE)
_MED = Side(style="medium", color=NAVY)
BOX = Border(top=_THIN, bottom=_THIN, left=_THIN, right=_THIN)
TOP_RULE = Border(top=_MED, bottom=_THIN, left=_THIN, right=_THIN)
BOTTOM_RULE = Border(top=_THIN, bottom=_MED, left=_THIN, right=_THIN)

# --- number formats --------------------------------------------------------
MONEY = '#,##0.00;[Red]-#,##0.00'
MONEY0 = '#,##0;[Red]-#,##0'
INT = '#,##0'
PERCENT = '0.0%'
PERCENT2 = '0.00%'
RATE = '#,##0.0000'
MULTIPLE = '0.00"x"'
SCHEME = '0.0"%"'
TEXT = '@'

LEFT = Alignment(horizontal="left", vertical="center")
CENTRE = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
INDENT = Alignment(horizontal="left", vertical="center", indent=1)


def put(ws, row: int, col: int, value=None, *, font=BODY, fill=None, fmt=None,
        align=None, border=BOX):
    """Write one fully-styled cell.  Styling is always explicit so nothing is
    inherited from the template underneath."""
    cell = ws.cell(row=row, column=col)
    cell.value = value
    cell.font = font
    cell.fill = fill if fill is not None else NO_FILL
    cell.border = border if border is not None else Border()
    if fmt:
        cell.number_format = fmt
    if align is not None:
        cell.alignment = align
    return cell


def wipe(ws, first_row: int, last_row: int, first_col: int, last_col: int) -> None:
    """Blank a rectangle completely — value and styling.

    The template ships pre-filled with dummy formulas; any cell left untouched
    would otherwise keep them and show stale numbers on a heading row.
    """
    for row in range(first_row, last_row + 1):
        for col in range(first_col, last_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.value = None
            cell.font = BODY
            cell.fill = NO_FILL
            cell.border = Border()
            cell.number_format = "General"
            cell.alignment = Alignment()


def unmerge(ws) -> None:
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))


def merge(ws, row: int, first_col: int, last_col: int) -> None:
    if last_col > first_col:
        ws.merge_cells(
            start_row=row, start_column=first_col, end_row=row, end_column=last_col
        )


def title(ws, row: int, text: str, last_col: int, first_col: int = 1) -> None:
    put(ws, row, first_col, text, font=TITLE, align=LEFT, border=None)
    merge(ws, row, first_col, last_col)
    ws.row_dimensions[row].height = 22


def subtitle(ws, row: int, text: str, last_col: int, first_col: int = 1) -> None:
    put(ws, row, first_col, text, font=SUBTITLE, align=WRAP, border=None)
    merge(ws, row, first_col, last_col)
    ws.row_dimensions[row].height = 26 if len(text) > 110 else 14


def section(ws, row: int, text: str, last_col: int, first_col: int = 1) -> None:
    """A merged, centred band that introduces a block of rows."""
    for col in range(first_col, last_col + 1):
        put(ws, row, col, None, fill=BAND_FILL, border=BOX)
    put(ws, row, first_col, text, font=SECTION, fill=BAND_FILL, align=CENTRE)
    merge(ws, row, first_col, last_col)
    ws.row_dimensions[row].height = 18


def table_title(ws, row: int, text: str, first_col: int, last_col: int) -> None:
    """A merged, centred caption sitting directly above a table's header row."""
    for col in range(first_col, last_col + 1):
        put(ws, row, col, None, fill=BAND_FILL, border=BOX)
    put(ws, row, first_col, text, font=SECTION, fill=BAND_FILL, align=CENTRE)
    merge(ws, row, first_col, last_col)
    ws.row_dimensions[row].height = 18


def header(ws, row: int, labels, first_col: int = 1, height: int = 30) -> int:
    """White bold text on navy, centred and wrapped.  Returns the column after."""
    col = first_col
    for label in labels:
        put(ws, row, col, label, font=HEADER, fill=HEADER_FILL,
            align=CENTRE, border=BOX)
        col += 1
    ws.row_dimensions[row].height = height
    return col


def note(ws, row: int, text: str, first_col: int, last_col: int) -> None:
    put(ws, row, first_col, text, font=NOTE, align=WRAP, border=None)
    merge(ws, row, first_col, last_col)
    # Roughly 105 characters fit per line across a full-width merge.
    lines = max(1, len(text) // 105 + 1)
    ws.row_dimensions[row].height = 13 * lines


def widths(ws, mapping: dict[str, float]) -> None:
    for column, width in mapping.items():
        ws.column_dimensions[column].width = width


def numeric_widths(ws, first_col: int, last_col: int, width: float = 13.0) -> None:
    for col in range(first_col, last_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def freeze(ws, row: int, col: int) -> None:
    # Use a coordinate string: the anchor may fall inside a merged range, and a
    # MergedCell cannot be assigned to freeze_panes.
    ws.freeze_panes = f"{get_column_letter(col)}{row}"
