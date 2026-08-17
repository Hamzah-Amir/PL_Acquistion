"""Write the finished workbook.

The user's template (`Amazon_UK_PL_TEMPLATE.xlsx`) is loaded and its input tabs
are filled, so the delivered file keeps the template's own structure and colour
key.  The analysis tabs listed in workflow §5 are then appended.

Every region is wiped before it is written: the template ships pre-filled with
dummy formulas, and any cell left untouched would keep them.

Colour key (workflow §2): blue = hard input from a source file, black = formula,
green = cross-sheet link, yellow = editable assumption.
"""
from __future__ import annotations

import os

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from . import styles as S
from .model import PLModel

TEMPLATE_ROWS = 13   # the template is laid out for 13 months
FIRST_MONTH_COL = 2  # column B on the P&L
LAST_TEMPLATE_COL = 32


class PL:
    """Row map for the P&L sheet, so the Summary tab can never drift from it."""

    TITLE = 1
    SUBTITLE = 2
    HEADER = 3
    REVENUE = 4
    PRODUCT_SALES = 5
    OTHER_INCOME = 6
    PROMO = 7
    RETURNS = 8
    NET_SALES = 9
    INCOME_CHECK = 10
    INCOME_DIFF = 11
    SALES_TAX = 12
    GROSS_SALES = 13

    FEES_HEAD = 15
    REFERRAL_EX = 16
    REFERRAL_VAT = 17
    FBA_EX = 18
    FBA_VAT = 19
    STORAGE = 20
    OTHER_FEES = 21
    TOTAL_FEES = 22

    ADS_HEAD = 24
    PPC_EX = 25
    PPC_VAT = 26
    TOTAL_PPC = 27
    EXPENSE_CHECK = 28
    EXPENSE_DIFF = 29

    COGS_HEAD = 31
    UNITS = 32
    ASP = 33
    COGS = 34
    INPUT_VAT = 35
    OUTPUT_VAT = 36
    OPEX = 37
    PROFIT = 38

    METRICS_HEAD = 40
    M_REFERRAL = 41
    M_FBA = 42
    M_STORAGE = 43
    M_OTHER = 44
    M_TOTAL_FEES = 45
    M_PPC = 46
    M_COGS = 47
    M_VAT = 48
    M_OPEX = 49
    M_MARGIN = 50
    M_TOTAL = 51
    M_INPUT_VAT = 52
    M_SALES_TAX = 53

    ASSUMPTIONS = 55
    LANDED = 56
    OPEX_INPUT = 57
    VAT_SCHEME = 58
    NOTES = 59
    LAST = 66


# ---------------------------------------------------------------------------
# Inputs_Monthly
# ---------------------------------------------------------------------------

MONTHLY_HEADERS = [
    "Month", "ProductSales FBA", "ProductSales SF", "Refund FBA", "Refund SF",
    "Delivery credit refund", "Postage credits", "Inventory credit",
    "Gift-wrap credits", "Promo rebates", "Promo rebate refunds", "INCOME TOTAL",
    "FBA selling fees", "SF selling fees", "Selling fee refunds",
    "FBA transaction fees", "FBA txn fee refunds", "Other txn fees",
    "Other txn fee refunds", "Storage & inbound", "Service fees", "Refund admin",
    "Delivery label", "Adjustments", "Cost of Advertising", "EXPENSES TOTAL",
    "Other income (residual)", "Other expenses (residual)", "Sales tax collected",
]


def _fill_inputs_monthly(ws, model: PLModel):
    """One row per month, straight off the Summary PDFs (workflow step 1)."""
    last_col = len(MONTHLY_HEADERS)
    S.wipe(ws, 1, TEMPLATE_ROWS + 6, 1, LAST_TEMPLATE_COL)

    S.header(ws, 1, MONTHLY_HEADERS, first_col=1, height=42)

    for index in range(len(model.months)):
        row = 2 + index
        month = model.months[index]
        source = model.settings["parse"]["months"].get(month.key, {})
        values = [
            source.get("product_sales_fba", 0.0),
            source.get("product_sales_sf", 0.0),
            source.get("refund_fba", 0.0),
            source.get("refund_sf", 0.0),
            source.get("delivery_credit_refunds", 0.0),
            source.get("postage_credits", 0.0),
            source.get("inventory_credit", 0.0),
            source.get("giftwrap_credits", 0.0),
            source.get("promo_rebates", 0.0),
            source.get("promo_rebate_refunds", 0.0),
            month.income_total,
            source.get("selling_fees_fba", 0.0),
            source.get("selling_fees_sf", 0.0),
            source.get("selling_fee_refunds", 0.0),
            source.get("fba_txn_fees", 0.0),
            source.get("fba_txn_fee_refunds", 0.0),
            source.get("other_txn_fees", 0.0),
            source.get("other_txn_fee_refunds", 0.0),
            source.get("storage_inbound", 0.0),
            source.get("service_fees", 0.0),
            source.get("refund_admin_fees", 0.0),
            source.get("delivery_label_purchases", 0.0),
            source.get("adjustments", 0.0),
            source.get("cost_of_advertising", 0.0),
            month.expenses_total,
            source.get("residual_income", 0.0),
            source.get("residual_expense", 0.0),
            month.sales_tax,
        ]
        stripe = S.ZEBRA_FILL if index % 2 else None
        S.put(ws, row, 1, month.label, font=S.BOLD, fill=stripe, align=S.LEFT)
        for offset, value in enumerate(values):
            # The two printed totals are the reconciliation anchors.
            emphasis = offset in (10, 24)
            S.put(ws, row, 2 + offset, value,
                  font=S.BOLD if emphasis else S.INPUT,
                  fill=S.TOTAL_FILL if emphasis else stripe,
                  fmt=S.MONEY, align=S.RIGHT)

    S.widths(ws, {"A": 11})
    S.numeric_widths(ws, 2, last_col, 13.5)
    S.freeze(ws, 2, 2)


# ---------------------------------------------------------------------------
# Inputs_Units
# ---------------------------------------------------------------------------

def _fill_inputs_units(ws, model: PLModel) -> tuple[int, int]:
    """Units per SKU per month plus the editable Cost Rate Table.

    Returns the (units total, COGS total) row numbers so the P&L can point at
    them — they move with the number of SKUs.
    """
    selling = [s for s in model.skus if any(s.units(m.key) for m in model.months)] \
        or model.skus
    months = len(model.months)
    first_row, last_row = 4, 4 + max(len(selling), 1) - 1
    units_row, cogs_row = last_row + 2, last_row + 3
    last_month_col = 3 + months
    rate_col, value_col = last_month_col + 2, last_month_col + 3

    S.wipe(ws, 1, cogs_row + 8, 1, max(LAST_TEMPLATE_COL, value_col + 2))

    S.title(ws, 1, "UNITS PER SKU PER MONTH", last_month_col)
    S.subtitle(
        ws, 2,
        "Net units (orders − refunds) from the transaction CSVs. Each SKU's "
        "£/unit points at its family in the COST RATE TABLE on the right — edit a "
        "rate there and every SKU on that family reprices.",
        last_month_col,
    )

    S.header(ws, 3, ["SKU", "Family", "£/unit"]
             + [m.label for m in model.months], first_col=1)

    vat_paid = model.settings.get("cogs_vat_paid", True)
    vat_included = model.settings.get("cogs_vat_included", True)
    grossing_up = vat_paid and not vat_included
    gross_up = "*1.2" if grossing_up else ""

    family_names = sorted(model.families) or ["Family 1"]
    family_row = {name: 4 + i for i, name in enumerate(family_names)}

    S.table_title(ws, 3, "COST RATE TABLE (edit)", rate_col, value_col)
    for name, row in family_row.items():
        S.put(ws, row, rate_col, name, font=S.BOLD, align=S.LEFT)
        S.put(ws, row, value_col, round(model.families.get(name, 0.0), 4),
              font=S.FORMULA, fill=S.YELLOW_FILL, fmt=S.RATE, align=S.RIGHT)
    if grossing_up:
        rate_note = ("Rates are EX-VAT; the £/unit column grosses them up by 20%.")
    elif vat_paid:
        rate_note = "Rates are VAT-INCLUSIVE."
    else:
        rate_note = "No VAT paid on goods — rates used exactly as entered."
    S.note(ws, 4 + len(family_names) + 1, rate_note, rate_col, value_col)

    for index, sku in enumerate(selling):
        row = first_row + index
        stripe = S.ZEBRA_FILL if index % 2 else None
        S.put(ws, row, 1, sku.sku, font=S.INPUT, fill=stripe, align=S.LEFT)
        S.put(ws, row, 2, sku.family, font=S.INPUT, fill=stripe, align=S.LEFT)
        target = family_row.get(sku.family)
        if target:
            S.put(ws, row, 3, f"=${get_column_letter(value_col)}${target}{gross_up}",
                  font=S.LINK, fill=stripe, fmt=S.RATE, align=S.RIGHT)
        else:
            S.put(ws, row, 3, round(sku.rate * (1.2 if grossing_up else 1.0), 4),
                  font=S.FORMULA, fill=stripe, fmt=S.RATE, align=S.RIGHT)
        for offset, month in enumerate(model.months):
            S.put(ws, row, 4 + offset, sku.units(month.key),
                  font=S.INPUT, fill=stripe, fmt=S.INT, align=S.RIGHT)

    for row, label, template in (
        (units_row, "MONTHLY UNITS (total)", "=SUM({c}{a}:{c}{b})"),
        (cogs_row, "MONTHLY COGS (Σ units × rate)",
         "=SUMPRODUCT($C${a}:$C${b},{c}{a}:{c}{b})"),
    ):
        S.put(ws, row, 1, label, font=S.TOTAL_FONT, fill=S.TOTAL_FILL,
              align=S.LEFT, border=S.TOP_RULE)
        for col in (2, 3):
            S.put(ws, row, col, None, fill=S.TOTAL_FILL, border=S.TOP_RULE)
        for offset in range(months):
            col = 4 + offset
            letter = get_column_letter(col)
            S.put(ws, row, col,
                  template.format(c=letter, a=first_row, b=last_row),
                  font=S.TOTAL_FONT, fill=S.TOTAL_FILL,
                  fmt=S.INT if row == units_row else S.MONEY,
                  align=S.RIGHT, border=S.TOP_RULE)

    S.widths(ws, {"A": 40, "B": 22, "C": 10})
    S.numeric_widths(ws, 4, last_month_col, 11.5)
    S.widths(ws, {get_column_letter(rate_col): 26,
                  get_column_letter(value_col): 12})
    S.freeze(ws, 4, 4)
    return units_row, cogs_row


# ---------------------------------------------------------------------------
# Inputs_Traffic
# ---------------------------------------------------------------------------

def _fill_inputs_traffic(ws, model: PLModel):
    S.wipe(ws, 1, TEMPLATE_ROWS + 6, 1, 10)
    S.header(ws, 1, ["Month", "Sessions", "Page views", "Units ordered",
                     "Buy Box %", "Child ASINs"])
    for index, month in enumerate(model.months):
        row = 2 + index
        stripe = S.ZEBRA_FILL if index % 2 else None
        S.put(ws, row, 1, month.label, font=S.BOLD, fill=stripe, align=S.LEFT)
        for offset, (value, fmt) in enumerate(
            [
                (month.sessions, S.INT),
                (month.page_views, S.INT),
                (month.units_ordered, S.INT),
                (month.buy_box, S.PERCENT),
                (month.child_asins, S.INT),
            ]
        ):
            S.put(ws, row, 2 + offset, value, font=S.INPUT, fill=stripe,
                  fmt=fmt, align=S.RIGHT)
    S.widths(ws, {"A": 11})
    S.numeric_widths(ws, 2, 6, 14)
    S.freeze(ws, 2, 2)


# ---------------------------------------------------------------------------
# P&L
# ---------------------------------------------------------------------------

def _fill_pl(ws, model: PLModel, units_row: int, cogs_row: int):
    """Regenerate the whole P&L grid for however many months were loaded."""
    count = len(model.months)
    last_col = FIRST_MONTH_COL + count - 1
    ttm_col = last_col + 1
    avg_col = last_col + 2
    # TTM drops the oldest month once a full 13 are present (workflow §2).
    ttm_start = FIRST_MONTH_COL + 1 if count > 12 else FIRST_MONTH_COL
    ttm_months = last_col - ttm_start + 1

    S.wipe(ws, 1, PL.LAST, 1, max(LAST_TEMPLATE_COL, avg_col + 1))

    def letter(col: int) -> str:
        return get_column_letter(col)

    def label(row: int, text: str, *, bold=False, indent=True):
        S.put(ws, row, 1, text,
              font=S.TOTAL_FONT if bold else S.BODY,
              fill=S.TOTAL_FILL if bold else None,
              align=S.INDENT if indent and not bold else S.LEFT)

    def across(row, formula, *, font=S.FORMULA, fmt=S.MONEY, fill=None):
        for index in range(count):
            col = FIRST_MONTH_COL + index
            S.put(ws, row, col, formula(index, col), font=font, fill=fill,
                  fmt=fmt, align=S.RIGHT)

    def totals(row, *, fmt=S.MONEY, avg=True, bold=False):
        font = S.TOTAL_FONT if bold else S.FORMULA
        S.put(ws, row, ttm_col,
              f"=SUM({letter(ttm_start)}{row}:{letter(last_col)}{row})",
              font=font, fill=S.TOTAL_FILL, fmt=fmt, align=S.RIGHT)
        if avg:
            S.put(ws, row, avg_col, f"={letter(ttm_col)}{row}/{ttm_months}",
                  font=font, fill=S.TOTAL_FILL, fmt=fmt, align=S.RIGHT)
        else:
            S.put(ws, row, avg_col, None, fill=S.TOTAL_FILL)

    def ratio(row, numerator, *, memo=False):
        across(row, lambda i, c: f"=IFERROR({numerator.format(c=letter(c))}"
                                 f"/{letter(c)}{PL.NET_SALES},0)",
               font=S.FORMULA, fmt=S.PERCENT)
        S.put(ws, row, ttm_col,
              f"=IFERROR({numerator.format(c=letter(ttm_col))}"
              f"/{letter(ttm_col)}{PL.NET_SALES},0)",
              font=S.TOTAL_FONT, fill=S.TOTAL_FILL, fmt=S.PERCENT, align=S.RIGHT)
        S.put(ws, row, avg_col, None, fill=S.TOTAL_FILL)

    S.title(ws, PL.TITLE, "MONTHLY P&L — computed from the Inputs_* tabs", avg_col)
    S.subtitle(
        ws, PL.SUBTITLE,
        f"Costs negative, revenue positive. TTM = the newest {ttm_months} months "
        f"({model.months[ttm_start - FIRST_MONTH_COL].label} to "
        f"{model.months[-1].label}). Blue = source data, black = formula, "
        f"green = cross-sheet link, yellow = editable assumption.",
        avg_col,
    )

    S.header(ws, PL.HEADER,
             ["Line item"] + [m.label for m in model.months] + ["TTM", "TTM avg/mo"])

    def im(col_letter: str, index: int) -> str:
        return f"Inputs_Monthly!{col_letter}{2 + index}"

    # --- revenue -----------------------------------------------------------
    S.section(ws, PL.REVENUE, "REVENUE", avg_col)
    label(PL.PRODUCT_SALES, "Product Sales (FBA + seller-fulfilled)")
    label(PL.OTHER_INCOME, "Other Income (postage, inventory, gift-wrap)")
    label(PL.PROMO, "less: Promotional Rebates (net)")
    label(PL.RETURNS, "less: Returns & Refunds")
    label(PL.NET_SALES, "Net Sales (ex-tax)", bold=True, indent=False)
    label(PL.INCOME_CHECK, "check: Amazon 'Income' total (PDF)")
    label(PL.INCOME_DIFF, "check: Difference (must be 0)")
    label(PL.SALES_TAX, "Sales tax collected from buyers")
    label(PL.GROSS_SALES, "Gross Sales (incl tax) — the VAT base",
          bold=True, indent=False)

    across(PL.PRODUCT_SALES, lambda i, c: f"={im('B', i)}+{im('C', i)}", font=S.LINK)
    across(PL.OTHER_INCOME,
           lambda i, c: f"={im('G', i)}+{im('H', i)}+{im('I', i)}+{im('AA', i)}",
           font=S.LINK)
    across(PL.PROMO, lambda i, c: f"={im('J', i)}+{im('K', i)}", font=S.LINK)
    across(PL.RETURNS,
           lambda i, c: f"={im('D', i)}+{im('E', i)}+{im('F', i)}", font=S.LINK)
    across(PL.NET_SALES,
           lambda i, c: f"=SUM({letter(c)}{PL.PRODUCT_SALES}:{letter(c)}{PL.RETURNS})",
           font=S.TOTAL_FONT, fill=S.TOTAL_FILL)
    across(PL.INCOME_CHECK, lambda i, c: f"={im('L', i)}", font=S.LINK)
    across(PL.INCOME_DIFF,
           lambda i, c: f"={letter(c)}{PL.NET_SALES}-{letter(c)}{PL.INCOME_CHECK}",
           font=S.TOTAL_FONT)
    across(PL.SALES_TAX, lambda i, c: f"={im('AC', i)}", font=S.LINK)
    across(PL.GROSS_SALES,
           lambda i, c: f"={letter(c)}{PL.NET_SALES}+{letter(c)}{PL.SALES_TAX}",
           font=S.TOTAL_FONT, fill=S.TOTAL_FILL)

    for row in (PL.PRODUCT_SALES, PL.OTHER_INCOME, PL.PROMO, PL.RETURNS,
                PL.SALES_TAX):
        totals(row)
    totals(PL.NET_SALES, bold=True)
    totals(PL.GROSS_SALES, bold=True)
    totals(PL.INCOME_CHECK, avg=False)
    S.put(ws, PL.INCOME_DIFF, ttm_col,
          f"={letter(ttm_col)}{PL.NET_SALES}-{letter(ttm_col)}{PL.INCOME_CHECK}",
          font=S.TOTAL_FONT, fill=S.TOTAL_FILL, fmt=S.MONEY, align=S.RIGHT)
    S.put(ws, PL.INCOME_DIFF, avg_col, None, fill=S.TOTAL_FILL)

    # --- Amazon fees -------------------------------------------------------
    S.section(ws, PL.FEES_HEAD,
              "AMAZON FEES  ·  VAT-inclusive; referral & FBA split net + VAT @20%",
              avg_col)
    label(PL.REFERRAL_EX, "Referral / Selling Fees (ex-VAT)")
    label(PL.REFERRAL_VAT, "Referral Fees — VAT @20%")
    label(PL.FBA_EX, "FBA Fulfilment Fees (ex-VAT)")
    label(PL.FBA_VAT, "FBA Fulfilment Fees — VAT @20%")
    label(PL.STORAGE, "Storage & Inbound Fees (incl VAT)")
    label(PL.OTHER_FEES, "Other Fees (incl VAT)")
    label(PL.TOTAL_FEES, "Total Amazon Fees (incl VAT, excl ads)",
          bold=True, indent=False)

    referral = lambda i: f"({im('M', i)}+{im('N', i)}+{im('O', i)})"
    fba = lambda i: f"({im('P', i)}+{im('Q', i)})"
    across(PL.REFERRAL_EX, lambda i, c: f"={referral(i)}/1.2", font=S.LINK)
    across(PL.REFERRAL_VAT,
           lambda i, c: f"={referral(i)}-{letter(c)}{PL.REFERRAL_EX}")
    across(PL.FBA_EX, lambda i, c: f"={fba(i)}/1.2", font=S.LINK)
    across(PL.FBA_VAT, lambda i, c: f"={fba(i)}-{letter(c)}{PL.FBA_EX}")
    across(PL.STORAGE, lambda i, c: f"={im('T', i)}", font=S.LINK)
    across(PL.OTHER_FEES,
           lambda i, c: "=" + "+".join(
               im(col, i) for col in ("R", "S", "U", "V", "W", "X", "AB")),
           font=S.LINK)
    across(PL.TOTAL_FEES,
           lambda i, c: f"=SUM({letter(c)}{PL.REFERRAL_EX}:{letter(c)}{PL.OTHER_FEES})",
           font=S.TOTAL_FONT, fill=S.TOTAL_FILL)
    for row in range(PL.REFERRAL_EX, PL.OTHER_FEES + 1):
        totals(row)
    totals(PL.TOTAL_FEES, bold=True)

    # --- advertising -------------------------------------------------------
    S.section(ws, PL.ADS_HEAD, "ADVERTISING / PPC", avg_col)
    label(PL.PPC_EX, "PPC / Advertising (ex-VAT)")
    label(PL.PPC_VAT, "PPC / Advertising — VAT @20%")
    label(PL.TOTAL_PPC, "Total PPC (incl VAT)", bold=True, indent=False)
    label(PL.EXPENSE_CHECK, "check: Amazon 'Expenses' total (PDF)")
    label(PL.EXPENSE_DIFF, "check: Difference (must be 0)")

    across(PL.PPC_EX, lambda i, c: f"={im('Y', i)}/1.2", font=S.LINK)
    across(PL.PPC_VAT, lambda i, c: f"={im('Y', i)}-{letter(c)}{PL.PPC_EX}")
    across(PL.TOTAL_PPC,
           lambda i, c: f"={letter(c)}{PL.PPC_EX}+{letter(c)}{PL.PPC_VAT}",
           font=S.TOTAL_FONT, fill=S.TOTAL_FILL)
    across(PL.EXPENSE_CHECK, lambda i, c: f"={im('Z', i)}", font=S.LINK)
    across(PL.EXPENSE_DIFF,
           lambda i, c: f"=({letter(c)}{PL.TOTAL_FEES}+{letter(c)}{PL.TOTAL_PPC})"
                        f"-{letter(c)}{PL.EXPENSE_CHECK}",
           font=S.TOTAL_FONT)
    totals(PL.PPC_EX)
    totals(PL.PPC_VAT)
    totals(PL.TOTAL_PPC, bold=True)
    totals(PL.EXPENSE_CHECK, avg=False)
    S.put(ws, PL.EXPENSE_DIFF, ttm_col,
          f"=({letter(ttm_col)}{PL.TOTAL_FEES}+{letter(ttm_col)}{PL.TOTAL_PPC})"
          f"-{letter(ttm_col)}{PL.EXPENSE_CHECK}",
          font=S.TOTAL_FONT, fill=S.TOTAL_FILL, fmt=S.MONEY, align=S.RIGHT)
    S.put(ws, PL.EXPENSE_DIFF, avg_col, None, fill=S.TOTAL_FILL)

    # --- COGS, VAT, profit -------------------------------------------------
    cogs_vat_paid = model.settings.get("cogs_vat_paid", True)
    S.section(ws, PL.COGS_HEAD, "COGS, VAT & NET PROFIT", avg_col)
    label(PL.UNITS, "Units Sold (net)")
    label(PL.ASP, "Avg Selling Price (incl tax)")
    label(PL.COGS, "COGS (landed cost × units)"
          + (", incl VAT" if cogs_vat_paid else ", no VAT paid"))
    label(PL.INPUT_VAT, "Input VAT in fees, PPC"
          + (" & COGS" if cogs_vat_paid else "") + "  (memo — see note)")
    label(PL.OUTPUT_VAT, "Output VAT (per scheme)")
    label(PL.OPEX, "Off-Amazon Operating Costs")
    label(PL.PROFIT, "ESTIMATED NET PROFIT", bold=True, indent=False)

    across(PL.UNITS,
           lambda i, c: f"=Inputs_Units!{get_column_letter(4 + i)}{units_row}",
           font=S.LINK, fmt=S.INT)
    across(PL.ASP,
           lambda i, c: f"=IFERROR({letter(c)}{PL.GROSS_SALES}/{letter(c)}{PL.UNITS},0)")
    across(PL.COGS,
           lambda i, c: f"=-Inputs_Units!{get_column_letter(4 + i)}{cogs_row}",
           font=S.LINK)
    cogs_term = f"+{{c}}{PL.COGS}/6" if cogs_vat_paid else ""
    across(PL.INPUT_VAT,
           lambda i, c: f"={letter(c)}{PL.REFERRAL_VAT}+{letter(c)}{PL.FBA_VAT}"
                        f"+{letter(c)}{PL.PPC_VAT}+{letter(c)}{PL.STORAGE}/6"
                        f"+{letter(c)}{PL.OTHER_FEES}/6"
                        + cogs_term.format(c=letter(c)))
    # Output VAT is charged on GROSS sales: dividing by 6 or 107.5 only extracts
    # the tax correctly from a VAT-inclusive figure.
    across(
        PL.OUTPUT_VAT,
        lambda i, c: (
            f'=IF($B${PL.VAT_SCHEME}=20,'
            f'-{letter(c)}{PL.GROSS_SALES}/6-{letter(c)}{PL.INPUT_VAT},'
            f'IF($B${PL.VAT_SCHEME}=7.5,'
            f'-{letter(c)}{PL.GROSS_SALES}/107.5*7.5,0))'
        ),
    )
    across(PL.OPEX, lambda i, c: f"=-$B${PL.OPEX_INPUT}")
    across(PL.PROFIT,
           lambda i, c: f"={letter(c)}{PL.NET_SALES}+{letter(c)}{PL.TOTAL_FEES}"
                        f"+{letter(c)}{PL.TOTAL_PPC}+{letter(c)}{PL.COGS}"
                        f"+{letter(c)}{PL.OUTPUT_VAT}+{letter(c)}{PL.OPEX}",
           font=S.TOTAL_FONT, fill=S.TOTAL_FILL)
    totals(PL.UNITS, fmt=S.INT)
    S.put(ws, PL.ASP, ttm_col,
          f"=IFERROR({letter(ttm_col)}{PL.GROSS_SALES}/{letter(ttm_col)}{PL.UNITS},0)",
          font=S.TOTAL_FONT, fill=S.TOTAL_FILL, fmt=S.MONEY, align=S.RIGHT)
    S.put(ws, PL.ASP, avg_col, None, fill=S.TOTAL_FILL)
    for row in (PL.COGS, PL.INPUT_VAT, PL.OUTPUT_VAT, PL.OPEX):
        totals(row)
    totals(PL.PROFIT, bold=True)

    # --- key metrics: a common-size P&L that adds to exactly 100% ----------
    S.section(ws, PL.METRICS_HEAD,
              "KEY METRICS  ·  % of Net Sales (the block adds to 100%)", avg_col)
    metrics = [
        (PL.M_REFERRAL, "Referral / Selling fee %",
         f"-({{c}}{PL.REFERRAL_EX}+{{c}}{PL.REFERRAL_VAT})"),
        (PL.M_FBA, "FBA fulfilment fee %", f"-({{c}}{PL.FBA_EX}+{{c}}{PL.FBA_VAT})"),
        (PL.M_STORAGE, "Storage & inbound %", f"-{{c}}{PL.STORAGE}"),
        (PL.M_OTHER, "Other fees %", f"-{{c}}{PL.OTHER_FEES}"),
        # Not prefixed with "=" — Excel would read the label as a formula.
        (PL.M_TOTAL_FEES, "Total Amazon fee % (subtotal)", f"-{{c}}{PL.TOTAL_FEES}"),
        (PL.M_PPC, "PPC / TACoS %", f"-{{c}}{PL.TOTAL_PPC}"),
        (PL.M_COGS, "COGS %", f"-{{c}}{PL.COGS}"),
        (PL.M_VAT, "Output VAT %", f"-{{c}}{PL.OUTPUT_VAT}"),
        (PL.M_OPEX, "Off-Amazon OpEx %", f"-{{c}}{PL.OPEX}"),
        (PL.M_MARGIN, "Net margin %", f"{{c}}{PL.PROFIT}"),
    ]
    for row, text, numerator in metrics:
        subtotal = row == PL.M_TOTAL_FEES
        S.put(ws, row, 1, text,
              font=S.BOLD if subtotal else S.BODY,
              fill=S.TOTAL_FILL if subtotal else None,
              align=S.LEFT if subtotal else S.INDENT)
        ratio(row, numerator)
        if subtotal:
            for col in range(FIRST_MONTH_COL, last_col + 1):
                ws.cell(row=row, column=col).fill = S.TOTAL_FILL
                ws.cell(row=row, column=col).font = S.BOLD

    # Row M_TOTAL_FEES is a subtotal of the four above it, so it is excluded.
    S.put(ws, PL.M_TOTAL, 1, "TOTAL (must be 100%)", font=S.TOTAL_FONT,
          fill=S.TOTAL_FILL, align=S.LEFT, border=S.TOP_RULE)
    total_formula = (
        f"=SUM({{c}}{PL.M_REFERRAL}:{{c}}{PL.M_OTHER})"
        f"+SUM({{c}}{PL.M_PPC}:{{c}}{PL.M_MARGIN})"
    )
    for index in range(count):
        col = FIRST_MONTH_COL + index
        S.put(ws, PL.M_TOTAL, col, total_formula.format(c=letter(col)),
              font=S.TOTAL_FONT, fill=S.TOTAL_FILL, fmt=S.PERCENT,
              align=S.RIGHT, border=S.TOP_RULE)
    S.put(ws, PL.M_TOTAL, ttm_col, total_formula.format(c=letter(ttm_col)),
          font=S.TOTAL_FONT, fill=S.TOTAL_FILL, fmt=S.PERCENT,
          align=S.RIGHT, border=S.TOP_RULE)
    S.put(ws, PL.M_TOTAL, avg_col, None, fill=S.TOTAL_FILL, border=S.TOP_RULE)

    S.put(ws, PL.M_INPUT_VAT, 1,
          "memo: Input VAT % (already inside the lines above)",
          font=S.NOTE, align=S.INDENT)
    ratio(PL.M_INPUT_VAT, f"-{{c}}{PL.INPUT_VAT}")
    S.put(ws, PL.M_SALES_TAX, 1, "memo: Sales tax % of net sales",
          font=S.NOTE, align=S.INDENT)
    ratio(PL.M_SALES_TAX, f"{{c}}{PL.SALES_TAX}")

    # --- assumptions -------------------------------------------------------
    S.section(ws, PL.ASSUMPTIONS, "ASSUMPTIONS  ·  yellow cells are editable",
              avg_col)
    S.put(ws, PL.LANDED, 1, "Blended landed £/unit (computed)",
          font=S.BOLD, align=S.LEFT)
    S.put(ws, PL.LANDED, 2,
          f"=IFERROR(-{letter(ttm_col)}{PL.COGS}/{letter(ttm_col)}{PL.UNITS},0)",
          font=S.FORMULA, fmt=S.RATE, align=S.RIGHT)
    S.put(ws, PL.OPEX_INPUT, 1, "Off-Amazon operating cost (£/month)",
          font=S.BOLD, align=S.LEFT)
    S.put(ws, PL.OPEX_INPUT, 2,
          round(float(model.settings.get("opex_monthly") or 0), 2),
          font=S.FORMULA, fill=S.YELLOW_FILL, fmt=S.MONEY, align=S.RIGHT)
    S.put(ws, PL.VAT_SCHEME, 1, "VAT scheme — enter 0, 7.5 or 20",
          font=S.BOLD, align=S.LEFT)
    S.put(ws, PL.VAT_SCHEME, 2, float(model.scheme),
          font=S.FORMULA, fill=S.YELLOW_FILL, fmt=S.SCHEME, align=S.RIGHT)

    if not cogs_vat_paid:
        cogs_note = ("No VAT was paid on the goods, so the landed cost is used "
                     "exactly as entered and contributes no input VAT.")
    elif model.settings.get("cogs_vat_included", True):
        cogs_note = "Landed cost entered VAT-INCLUSIVE (import VAT = COGS / 6)."
    else:
        cogs_note = ("Landed cost entered EX-VAT, grossed up by 20% "
                     "(import VAT = rate × 20%).")
    notes = [
        f"Currently: {model.scheme_label}.  {cogs_note}",
        f"Output VAT is charged on GROSS sales (row {PL.GROSS_SALES} = Net Sales + "
        "sales tax collected), because dividing by 6 or by 107.5 only extracts the "
        "tax correctly from a VAT-inclusive figure.",
        "0% = not registered, no output VAT.   7.5% = flat rate, Gross Sales / 107.5 "
        "× 7.5.   20% = standard, Gross Sales / 6 less reclaimable input VAT — a "
        "positive figure there means a repayment is due from HMRC.",
        f"On the 0% and 7.5% schemes the input VAT on row {PL.INPUT_VAT} is "
        "IRRECOVERABLE. It already sits inside the VAT-inclusive fee, PPC and COGS "
        "lines, so it is a memo only and is deliberately NOT deducted again — that "
        "would double-count it.",
        f"Net Sales (row {PL.NET_SALES}) reconciles to Amazon's printed Income "
        f"total; sales tax is reported separately on row {PL.SALES_TAX} because "
        "Amazon totals it in its own block.",
    ]
    for offset, text in enumerate(notes):
        S.note(ws, PL.NOTES + offset, text, 1, avg_col)

    S.widths(ws, {"A": 42})
    S.numeric_widths(ws, FIRST_MONTH_COL, avg_col, 13.0)
    S.freeze(ws, PL.REVENUE, FIRST_MONTH_COL)


# ---------------------------------------------------------------------------
# Summary
# ---------------------------------------------------------------------------

def _fill_summary(ws, model: PLModel):
    S.wipe(ws, 1, 40, 1, 10)
    S.widths(ws, {"A": 3, "B": 40, "C": 16, "D": 3, "E": 34, "F": 16})

    S.title(ws, 2, "SUMMARY — trailing twelve months", 6, first_col=2)
    S.subtitle(
        ws, 3,
        f"{model.ttm[0].label} to {model.ttm[-1].label}.  "
        f"VAT scheme: {model.scheme_label}.", 6, first_col=2,
    )

    rows = [
        ("Net Sales (ex-tax)", f"O{PL.NET_SALES}", S.MONEY, False),
        ("Sales tax collected", f"O{PL.SALES_TAX}", S.MONEY, False),
        ("Gross Sales (incl tax)", f"O{PL.GROSS_SALES}", S.MONEY, True),
        ("Units sold (net)", f"O{PL.UNITS}", S.INT, False),
        ("Avg selling price (incl tax)", f"O{PL.ASP}", S.MONEY, False),
        ("Total Amazon Fees incl VAT", f"O{PL.TOTAL_FEES}", S.MONEY, False),
        ("PPC / Advertising incl VAT", f"O{PL.TOTAL_PPC}", S.MONEY, False),
        ("COGS", f"O{PL.COGS}", S.MONEY, False),
        (
            "Input VAT in fees, PPC & COGS"
            + (" (reclaimed)" if model.input_vat_is_reclaimable else " (irrecoverable)"),
            f"O{PL.INPUT_VAT}", S.MONEY, False,
        ),
        ("Output VAT (per scheme)", f"O{PL.OUTPUT_VAT}", S.MONEY, False),
        ("Off-Amazon OpEx", f"O{PL.OPEX}", S.MONEY, False),
        ("NET PROFIT (TTM)", f"O{PL.PROFIT}", S.MONEY, True),
    ]
    row = 5
    S.header(ws, row, ["Line", "TTM"], first_col=2, height=18)
    row += 1
    for index, (text, cell, fmt, emphasis) in enumerate(rows):
        stripe = S.ZEBRA_FILL if index % 2 else None
        S.put(ws, row, 2, text,
              font=S.TOTAL_FONT if emphasis else S.BODY,
              fill=S.TOTAL_FILL if emphasis else stripe, align=S.LEFT)
        S.put(ws, row, 3, f"='P&L'!{cell}",
              font=S.TOTAL_FONT if emphasis else S.LINK,
              fill=S.TOTAL_FILL if emphasis else stripe, fmt=fmt, align=S.RIGHT)
        row += 1

    row += 1
    S.header(ws, row, ["Ratio", "TTM"], first_col=2, height=18)
    row += 1
    ratios = [
        ("Net margin %", f"O{PL.M_MARGIN}", S.PERCENT),
        ("TACoS % (PPC / net sales)", f"O{PL.M_PPC}", S.PERCENT),
        ("Referral / Selling fee %", f"O{PL.M_REFERRAL}", S.PERCENT),
        ("FBA fulfilment fee %", f"O{PL.M_FBA}", S.PERCENT),
        ("Total Amazon fee %", f"O{PL.M_TOTAL_FEES}", S.PERCENT),
        ("COGS %", f"O{PL.M_COGS}", S.PERCENT),
        ("Blended landed £/unit", f"B{PL.LANDED}", S.RATE),
    ]
    for index, (text, cell, fmt) in enumerate(ratios):
        stripe = S.ZEBRA_FILL if index % 2 else None
        S.put(ws, row, 2, text, fill=stripe, align=S.LEFT)
        S.put(ws, row, 3, f"='P&L'!{cell}", font=S.LINK, fill=stripe,
              fmt=fmt, align=S.RIGHT)
        row += 1

    # --- valuation (right-hand column) ------------------------------------
    S.table_title(ws, 5, "VALUATION", 5, 6)
    S.put(ws, 6, 5, "Asking Price (enter)", font=S.BOLD, align=S.LEFT)
    S.put(ws, 6, 6, round(float(model.settings.get("asking_price") or 0), 2),
          font=S.FORMULA, fill=S.YELLOW_FILL, fmt=S.MONEY, align=S.RIGHT)
    S.put(ws, 7, 5, "Implied multiple (Price ÷ Net Profit)", font=S.BOLD,
          align=S.LEFT)
    S.put(ws, 7, 6, f"=IFERROR(F6/'P&L'!O{PL.PROFIT},0)",
          font=S.FORMULA, fmt=S.MULTIPLE, align=S.RIGHT)

    S.table_title(ws, 9, "SENSITIVITY — TTM net profit by landed cost", 5, 6)
    S.header(ws, 10, ["Landed £/unit", "Net profit"], first_col=5, height=18)
    base = model.blended_landed_cost or 1.0
    rates = sorted({round(max(base, 0.25) * m, 2) for m in (0.5, 0.75, 1.0, 1.5, 2.0)})
    for index, rate in enumerate(rates):
        line = 11 + index
        stripe = S.ZEBRA_FILL if index % 2 else None
        S.put(ws, line, 5, rate, fill=stripe, fmt=S.RATE, align=S.RIGHT)
        S.put(ws, line, 6,
              f"='P&L'!O{PL.PROFIT}-'P&L'!O{PL.COGS}-'P&L'!O{PL.UNITS}*E{line}",
              font=S.FORMULA, fill=stripe, fmt=S.MONEY, align=S.RIGHT)
    S.note(ws, 11 + len(rates) + 1,
           "COGS is the one figure that comes from you rather than from Amazon's "
           "exports — these rows show the deal at other landed costs.", 5, 6)
    S.freeze(ws, 6, 1)


# ---------------------------------------------------------------------------
# Analysis tabs (workflow §5)
# ---------------------------------------------------------------------------

def _sheet(wb, name: str):
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)


def _write_status(wb, model: PLModel):
    ws = _sheet(wb, "Status & Missing")
    S.widths(ws, {"A": 2, "B": 150})
    parse = model.settings["parse"]
    counts = parse["counts"]
    seller = model.seller or {}
    name = (model.settings.get("business_name") or seller.get("legal_name")
            or "This account")

    S.title(ws, 2, "DATA STATUS — what's loaded, what's missing, key flags", 2,
            first_col=2)
    S.subtitle(
        ws, 3,
        f"{name}  ·  {seller.get('display_name', '')}  ·  {len(model.months)} months "
        f"({model.months[0].label} – {model.months[-1].label})", 2, first_col=2,
    )

    row = 5

    def block(heading: str, lines: list[str]):
        nonlocal row
        S.table_title(ws, row, heading, 2, 2)
        row += 1
        for index, text in enumerate(lines):
            S.put(ws, row, 2, text, align=S.WRAP,
                  fill=S.ZEBRA_FILL if index % 2 else None)
            ws.row_dimensions[row].height = 13 * max(1, len(text) // 140 + 1)
            row += 1
        row += 1

    block("LOADED", [
        f"{counts['summary']} monthly Summary PDFs — the authoritative source; "
        "every revenue and fee line on the P&L comes from these.",
        f"{counts['transaction']} Transaction CSVs — units per SKU per month "
        f"({model.ttm_units:,} net units across the TTM) plus the fee cross-checks.",
        f"{counts['business']} Business Reports — sessions, page views, conversion "
        "and Buy Box; see 'Traffic & Conversion'.",
        f"{len(model.invoices)} advertising invoices "
        f"({parse.get('invoice_duplicates', 0)} duplicates removed); see 'PPC Invoices'.",
        f"{len(model.skus)} SKUs across {len(model.families)} product families; see "
        "'SKU Concentration' and 'COGS Build'.",
    ])

    block("RECONCILIATION  (workflow §4)", [
        ("PASS  " if check.passed else "WARN  " if check.severity == "warning"
         else "FAIL  ") + f"{check.name} — {check.detail}"
        for check in model.checks
    ])

    opex = float(model.settings.get("opex_monthly") or 0)
    price = float(model.settings.get("asking_price") or 0)
    vat_notes = {
        "0": "0% — not registered. Output VAT is £0. The "
             f"£{-model.ttm_sum('input_vat'):,.0f} of input VAT inside the fees, PPC "
             "and COGS is irrecoverable; it already sits in those VAT-inclusive cost "
             "lines and is reported as a memo, not deducted twice.",
        "7.5": "7.5% flat rate — output VAT is Gross Sales / 107.5 × 7.5, i.e. "
               f"£{-model.ttm_sum('output_vat'):,.0f} across the TTM. The "
               f"£{-model.ttm_sum('input_vat'):,.0f} of input VAT is irrecoverable and "
               "stays a cost inside the fee, PPC and COGS lines.",
        "20": "20% standard — output VAT is Gross Sales / 6 less reclaimable input "
              f"VAT: £{model.ttm_sum('gross_sales') / 6:,.0f} less "
              f"£{-model.ttm_sum('input_vat'):,.0f} = "
              f"£{-model.ttm_sum('output_vat'):,.0f} net across the TTM.",
    }
    cogs_basis = (
        "no VAT paid on the goods" if not model.settings.get("cogs_vat_paid", True)
        else "VAT-inclusive (import VAT = COGS / 6)"
        if model.settings.get("cogs_vat_included", True)
        else "ex-VAT, grossed up by 20% (import VAT = rate × 20%)"
    )
    block("ASSUMPTIONS APPLIED  (workflow step 8)", [
        "VAT: " + vat_notes.get(model.scheme, model.scheme_label),
        f"Landed cost: bottom-up from the Cost Rate Table, entered {cogs_basis}. "
        f"Blended £{model.blended_landed_cost:,.2f}/unit across "
        f"{model.ttm_units:,} TTM units.",
        f"Off-Amazon operating costs: £{opex:,.2f}/month."
        if opex else "Off-Amazon operating costs: none included.",
        f"Asking price: £{price:,.2f} — see the implied multiple on the Summary tab."
        if price else "Asking price: not supplied, so no valuation multiple is shown.",
    ])

    ttm_sales = model.ttm_sum("net_sales")
    flags: list[str] = []
    if ttm_sales > 90000 and model.scheme == "0":
        flags.append(
            f"TTM turnover £{ttm_sales:,.0f} is above the £90k UK VAT registration "
            "threshold but the model is set to 0% (not registered). A buyer will very "
            f"likely have to register — roughly £{model.ttm_sum('gross_sales') / 6:,.0f}"
            "/yr of output VAT. Model the 20% scheme before pricing the deal."
        )
    repayments = [m.label for m in model.months if m.output_vat > 0]
    if repayments:
        flags.append(
            "Reclaimable input VAT exceeds output VAT in "
            + ", ".join(repayments)
            + " — those months show a VAT repayment due from HMRC, not a bill."
        )
    unpriced = [s.sku for s in model.skus if s.rate <= 0 and s.units_by_month]
    if unpriced:
        flags.append(
            f"{len(unpriced)} SKU(s) carry a £0 landed cost, so their COGS is "
            "understated: " + ", ".join(unpriced[:8])
        )
    low_buy_box = [m for m in model.months if m.has_traffic and m.buy_box < 0.95]
    if low_buy_box:
        flags.append(
            "Buy Box dropped below 95% in "
            + ", ".join(f"{m.label} ({m.buy_box:.0%})" for m in low_buy_box)
            + " — check for a hijacker, pricing suppression or a stock-out."
        )
    flags += [
        "Confirm what actually transfers: the brand and ASINs only (asset sale), or "
        "the Seller Central account itself.",
        "Independent verification still needed: Account Health, trademark / Brand "
        "Registry, supplier invoices, and current inventory on hand.",
    ]
    flags += [w for w in model.warnings[:8]]
    block("FLAGS & THINGS STILL TO VERIFY", flags)


def _write_sku_concentration(wb, model: PLModel):
    ws = _sheet(wb, "SKU Concentration")
    S.wipe(ws, 1, 60, 1, 8)
    S.widths(ws, {"A": 2, "B": 42, "C": 16, "D": 12, "E": 26})

    ttm_keys = {m.key for m in model.ttm}
    rows = sorted(
        ((s.sku, round(sum(v for k, v in s.sales_by_month.items() if k in ttm_keys), 2),
          s.family) for s in model.skus),
        key=lambda r: r[1], reverse=True,
    )

    S.title(ws, 2, "SKU CONCENTRATION — trailing twelve months", 5, first_col=2)
    S.subtitle(ws, 3,
               "Revenue is gross of sales tax, as the buyer paid it. Variation SKUs "
               "(colours, sizes, refills) of one product share supply-chain and "
               "demand risk — read the family table, not just the SKU count.",
               5, first_col=2)

    S.header(ws, 5, ["SKU", "TTM Sales", "% of total", "Product family"], first_col=2)
    first = 6
    for index, (sku, sales, family) in enumerate(rows):
        row = first + index
        stripe = S.ZEBRA_FILL if index % 2 else None
        S.put(ws, row, 2, sku, fill=stripe, align=S.LEFT)
        S.put(ws, row, 3, sales, fill=stripe, fmt=S.MONEY, align=S.RIGHT)
        S.put(ws, row, 4, f"=IFERROR(C{row}/$C${first + len(rows)},0)",
              fill=stripe, fmt=S.PERCENT, align=S.RIGHT)
        S.put(ws, row, 5, family, fill=stripe, align=S.LEFT)
    total_row = first + len(rows)
    S.put(ws, total_row, 2, "TOTAL", font=S.TOTAL_FONT, fill=S.TOTAL_FILL,
          align=S.LEFT, border=S.TOP_RULE)
    S.put(ws, total_row, 3, f"=SUM(C{first}:C{total_row - 1})", font=S.TOTAL_FONT,
          fill=S.TOTAL_FILL, fmt=S.MONEY, align=S.RIGHT, border=S.TOP_RULE)
    for col in (4, 5):
        S.put(ws, total_row, col, None, fill=S.TOTAL_FILL, border=S.TOP_RULE)

    row = total_row + 2
    family_totals: dict[str, float] = {}
    for sku, sales, family in rows:
        family_totals[family] = round(family_totals.get(family, 0.0) + sales, 2)
    S.table_title(ws, row, "BY PRODUCT FAMILY", 2, 4)
    row += 1
    S.header(ws, row, ["Family", "TTM Sales", "% of total"], first_col=2)
    row += 1
    for index, (family, sales) in enumerate(
        sorted(family_totals.items(), key=lambda kv: kv[1], reverse=True)
    ):
        stripe = S.ZEBRA_FILL if index % 2 else None
        S.put(ws, row, 2, family, fill=stripe, align=S.LEFT)
        S.put(ws, row, 3, sales, fill=stripe, fmt=S.MONEY, align=S.RIGHT)
        S.put(ws, row, 4, f"=IFERROR(C{row}/$C${total_row},0)", fill=stripe,
              fmt=S.PERCENT, align=S.RIGHT)
        row += 1

    row += 1
    S.table_title(ws, row, "CONCENTRATION", 2, 3)
    row += 1
    last_sku = max(first, total_row - 1)
    for label, value, fmt in [
        ("Distinct SKUs", len(rows), S.INT),
        ("Top SKU %", f"=IFERROR(C{first}/$C${total_row},0)", S.PERCENT),
        ("Top 3 SKUs %",
         f"=IFERROR(SUM(C{first}:C{min(first + 2, last_sku)})/$C${total_row},0)",
         S.PERCENT),
        ("Top 5 SKUs %",
         f"=IFERROR(SUM(C{first}:C{min(first + 4, last_sku)})/$C${total_row},0)",
         S.PERCENT),
    ]:
        S.put(ws, row, 2, label, align=S.LEFT)
        S.put(ws, row, 3, value, font=S.TOTAL_FONT, fmt=fmt, align=S.RIGHT)
        row += 1
    S.freeze(ws, 6, 1)


def _write_cogs_build(wb, model: PLModel):
    ws = _sheet(wb, "COGS Build")
    months = len(model.months)
    ttm_units_col = 3 + months
    ttm_cogs_col = ttm_units_col + 1
    family_col = ttm_cogs_col + 1
    S.wipe(ws, 1, 60, 1, family_col + 2)

    S.title(ws, 1, "COGS BUILD — units per SKU × landed cost", family_col)
    S.subtitle(
        ws, 2,
        "Net units (orders − refunds) per SKU per month. £/unit is driven by the "
        "COST RATE TABLE on Inputs_Units; monthly COGS = Σ(units × cost). Compare the "
        "blended rate against whatever the seller quotes — a blanket figure weighted "
        "by the real sales mix is usually very different.",
        family_col,
    )

    S.header(ws, 3, ["SKU", "£/unit"] + [m.label for m in model.months]
             + ["TTM units", "TTM COGS", "Family"], first_col=1)

    selling = [s for s in model.skus if any(s.units(m.key) for m in model.months)] \
        or model.skus
    first = 4
    ttm_offset = months - len(model.ttm)
    for index, sku in enumerate(selling):
        row = first + index
        stripe = S.ZEBRA_FILL if index % 2 else None
        S.put(ws, row, 1, sku.sku, fill=stripe, align=S.LEFT)
        S.put(ws, row, 2, sku.rate, font=S.LINK, fill=stripe, fmt=S.RATE,
              align=S.RIGHT)
        for offset, month in enumerate(model.months):
            S.put(ws, row, 3 + offset, sku.units(month.key), fill=stripe,
                  fmt=S.INT, align=S.RIGHT)
        span = (f"{get_column_letter(3 + ttm_offset)}{row}:"
                f"{get_column_letter(ttm_units_col - 1)}{row}")
        S.put(ws, row, ttm_units_col, f"=SUM({span})", font=S.TOTAL_FONT,
              fill=stripe, fmt=S.INT, align=S.RIGHT)
        S.put(ws, row, ttm_cogs_col,
              f"={get_column_letter(ttm_units_col)}{row}*B{row}",
              font=S.TOTAL_FONT, fill=stripe, fmt=S.MONEY, align=S.RIGHT)
        S.put(ws, row, family_col, sku.family, fill=stripe, align=S.LEFT)

    last = first + len(selling) - 1
    cogs_row = last + 1
    units_row = last + 2
    for row, label, template, fmt in (
        (cogs_row, "MONTHLY COGS", "=SUMPRODUCT($B${a}:$B${b},{c}{a}:{c}{b})", S.MONEY),
        (units_row, "MONTHLY UNITS", "=SUM({c}{a}:{c}{b})", S.INT),
    ):
        S.put(ws, row, 1, label, font=S.TOTAL_FONT, fill=S.TOTAL_FILL,
              align=S.LEFT, border=S.TOP_RULE)
        S.put(ws, row, 2, None, fill=S.TOTAL_FILL, border=S.TOP_RULE)
        for offset in range(months):
            col = 3 + offset
            S.put(ws, row, col,
                  template.format(c=get_column_letter(col), a=first, b=last),
                  font=S.TOTAL_FONT, fill=S.TOTAL_FILL, fmt=fmt,
                  align=S.RIGHT, border=S.TOP_RULE)
        S.put(ws, row, ttm_units_col,
              f"=SUM({get_column_letter(3 + ttm_offset)}{row}:"
              f"{get_column_letter(ttm_units_col - 1)}{row})",
              font=S.TOTAL_FONT, fill=S.TOTAL_FILL, fmt=fmt,
              align=S.RIGHT, border=S.TOP_RULE)
        for col in (ttm_cogs_col, family_col):
            S.put(ws, row, col, None, fill=S.TOTAL_FILL, border=S.TOP_RULE)

    row = units_row + 2
    S.put(ws, row, 1, "Blended landed £/unit (TTM)", font=S.BOLD, align=S.LEFT)
    S.put(ws, row, 2,
          f"=IFERROR({get_column_letter(ttm_units_col)}{cogs_row}"
          f"/{get_column_letter(ttm_units_col)}{units_row},0)",
          font=S.TOTAL_FONT, fmt=S.RATE, align=S.RIGHT)

    S.widths(ws, {"A": 42, "B": 10})
    S.numeric_widths(ws, 3, ttm_units_col, 11.5)
    S.widths(ws, {get_column_letter(ttm_cogs_col): 14,
                  get_column_letter(family_col): 24})
    S.freeze(ws, 4, 3)


def _write_ppc(wb, model: PLModel):
    ws = _sheet(wb, "PPC Invoices")
    S.wipe(ws, 1, 40 + len(model.invoices) + 12, 1, 10)
    S.widths(ws, {"A": 2, "B": 22, "C": 9, "D": 14, "E": 14, "F": 14, "G": 14, "H": 14})

    S.title(ws, 2, f"PPC INVOICES — {len(model.invoices)} advertising tax invoices",
            8, first_col=2)
    S.subtitle(
        ws, 3,
        "VAT = gross − net (each invoice is net × 1.20). Bucketed by invoice date and "
        "compared to 'Cost of Advertising' on the settlement. The P&L uses the "
        "settlement figure — that is what reconciles to the Expenses total.",
        8, first_col=2,
    )

    S.table_title(ws, 5, "MONTHLY SUMMARY (bucketed by invoice date)", 2, 8)
    S.header(ws, 6, ["Month", "# inv", "PPC net", "PPC VAT", "PPC gross",
                     "Settlement", "Diff"], first_col=2)
    first = 7
    for index, month in enumerate(model.months):
        row = first + index
        stripe = S.ZEBRA_FILL if index % 2 else None
        S.put(ws, row, 2, month.label, fill=stripe, align=S.LEFT)
        S.put(ws, row, 3, month.ppc_invoice_count, fill=stripe, fmt=S.INT,
              align=S.RIGHT)
        S.put(ws, row, 4, month.ppc_invoice_net, fill=stripe, fmt=S.MONEY,
              align=S.RIGHT)
        S.put(ws, row, 5, round(month.ppc_invoice_gross - month.ppc_invoice_net, 2),
              fill=stripe, fmt=S.MONEY, align=S.RIGHT)
        S.put(ws, row, 6, f"=D{row}+E{row}", font=S.TOTAL_FONT, fill=stripe,
              fmt=S.MONEY, align=S.RIGHT)
        S.put(ws, row, 7, -month.total_ppc, fill=stripe, fmt=S.MONEY, align=S.RIGHT)
        S.put(ws, row, 8, f"=F{row}-G{row}", fill=stripe, fmt=S.MONEY, align=S.RIGHT)
    total_row = first + len(model.months)
    S.put(ws, total_row, 2, "TOTAL", font=S.TOTAL_FONT, fill=S.TOTAL_FILL,
          align=S.LEFT, border=S.TOP_RULE)
    for col, letter in ((3, "C"), (4, "D"), (5, "E"), (6, "F"), (7, "G")):
        S.put(ws, total_row, col, f"=SUM({letter}{first}:{letter}{total_row - 1})",
              font=S.TOTAL_FONT, fill=S.TOTAL_FILL,
              fmt=S.INT if col == 3 else S.MONEY, align=S.RIGHT, border=S.TOP_RULE)
    S.put(ws, total_row, 8, f"=F{total_row}-G{total_row}", font=S.TOTAL_FONT,
          fill=S.TOTAL_FILL, fmt=S.MONEY, align=S.RIGHT, border=S.TOP_RULE)

    matched = sum(1 for m in model.months if m.ppc_invoice_count
                  and abs(m.ppc_invoice_gross + m.total_ppc) <= 0.02)
    with_invoices = sum(1 for m in model.months if m.ppc_invoice_count)
    row = total_row + 2
    S.note(ws, row,
           f"Reconciliation: {matched} of {with_invoices} months with invoices tie to "
           "the settlement advertising to the penny. Differences are billing-cycle "
           "timing at month boundaries — an invoice raised on the 1st or 2nd covers "
           "spend from the previous month.", 2, 8)

    row += 2
    S.table_title(ws, row, f"ALL {len(model.invoices)} INVOICES", 2, 7)
    row += 1
    S.header(ws, row, ["Invoice #", "Date", "Period", "Net", "VAT", "Gross"],
             first_col=2)
    S.widths(ws, {"D": 24})
    detail_first = row + 1
    for index, invoice in enumerate(model.invoices):
        line = detail_first + index
        stripe = S.ZEBRA_FILL if index % 2 else None
        S.put(ws, line, 2, invoice["number"], fill=stripe, align=S.LEFT)
        S.put(ws, line, 3, invoice["date_display"], fill=stripe, align=S.CENTRE)
        S.put(ws, line, 4, invoice["period"], fill=stripe, align=S.CENTRE)
        S.put(ws, line, 5, invoice["net"], fill=stripe, fmt=S.MONEY, align=S.RIGHT)
        S.put(ws, line, 6, invoice["vat"], fill=stripe, fmt=S.MONEY, align=S.RIGHT)
        S.put(ws, line, 7, invoice["gross"], fill=stripe, fmt=S.MONEY, align=S.RIGHT)
    if model.invoices:
        line = detail_first + len(model.invoices)
        S.put(ws, line, 2, "TOTAL", font=S.TOTAL_FONT, fill=S.TOTAL_FILL,
              align=S.LEFT, border=S.TOP_RULE)
        for col in (3, 4):
            S.put(ws, line, col, None, fill=S.TOTAL_FILL, border=S.TOP_RULE)
        for col, letter in ((5, "E"), (6, "F"), (7, "G")):
            S.put(ws, line, col,
                  f"=SUM({letter}{detail_first}:{letter}{line - 1})",
                  font=S.TOTAL_FONT, fill=S.TOTAL_FILL, fmt=S.MONEY,
                  align=S.RIGHT, border=S.TOP_RULE)
    S.freeze(ws, 7, 1)


def _write_fee_reconciliation(wb, model: PLModel):
    ws = _sheet(wb, "Fee Reconciliation")
    S.wipe(ws, 1, 70, 1, 16)
    S.widths(ws, {"A": 2, "B": 12})
    S.numeric_widths(ws, 3, 15, 13.5)

    S.title(ws, 2, "FEE RECONCILIATION — Summary PDF vs transaction CSV", 11,
            first_col=2)
    S.subtitle(
        ws, 3,
        "Confirms every Amazon fee on the P&L comes from the Payment report and ties "
        "out both ways. Costs negative; 'd' = CSV − Summary, so 0 is a tie. Referral "
        "is the order-line selling fee — the Summary's separate 'selling fee refunds' "
        "credit is its own line on the P&L.",
        11, first_col=2,
    )

    S.table_title(ws, 5, "REFERRAL (SELLING FEE)", 3, 5)
    S.table_title(ws, 5, "FBA FULFILMENT", 6, 8)
    S.table_title(ws, 5, "STORAGE & INBOUND", 9, 11)
    S.header(ws, 6, ["Month", "Summary", "CSV", "d", "Summary", "CSV", "d",
                     "Summary", "CSV", "d"], first_col=2)

    first = 7
    for index, month in enumerate(model.months):
        row = first + index
        stripe = S.ZEBRA_FILL if index % 2 else None
        S.put(ws, row, 2, month.label, fill=stripe, align=S.LEFT)
        for base, (summary, csv) in enumerate([
            (month.summary_referral, month.csv_referral),
            (month.summary_fba, month.csv_fba),
            (month.summary_storage, month.csv_storage),
        ]):
            col = 3 + base * 3
            S.put(ws, row, col, summary, fill=stripe, fmt=S.MONEY, align=S.RIGHT)
            S.put(ws, row, col + 1, csv, fill=stripe, fmt=S.MONEY, align=S.RIGHT)
            S.put(ws, row, col + 2,
                  f"={get_column_letter(col + 1)}{row}-{get_column_letter(col)}{row}",
                  font=S.TOTAL_FONT, fill=stripe, fmt=S.MONEY, align=S.RIGHT)
    total_row = first + len(model.months)
    S.put(ws, total_row, 2, "TOTAL", font=S.TOTAL_FONT, fill=S.TOTAL_FILL,
          align=S.LEFT, border=S.TOP_RULE)
    for col in range(3, 12):
        letter = get_column_letter(col)
        S.put(ws, total_row, col, f"=SUM({letter}{first}:{letter}{total_row - 1})",
              font=S.TOTAL_FONT, fill=S.TOTAL_FILL, fmt=S.MONEY,
              align=S.RIGHT, border=S.TOP_RULE)

    row = total_row + 2
    S.table_title(ws, row, "OTHER FEES — FULL COMPOSITION", 2, 15)
    row += 1
    S.note(ws, row,
           "Amazon bundles the monthly seller subscription, Deal fees and Coupon fees "
           "into one 'Service fees' line; they are split out here from the transaction "
           "data. Deal fees mark the months a paid Deal ran. The Total column ties to "
           "the P&L Other Fees line.", 2, 15)
    row += 1
    S.header(ws, row, ["Month", "Other txn fees", "Other txn refunds", "Subscription",
                       "Deal fees", "Coupon fees", "Refund admin", "Delivery labels",
                       "Adjustments", "Other", "Total Other", "P&L line", "d"],
             first_col=2)
    detail_first = row + 1
    for index, month in enumerate(model.months):
        line = detail_first + index
        stripe = S.ZEBRA_FILL if index % 2 else None
        S.put(ws, line, 2, month.label, fill=stripe, align=S.LEFT)
        for offset, value in enumerate([
            month.other_txn_fees, month.other_txn_fee_refunds,
            month.subscription_fees, month.deal_fees, month.coupon_fees,
            month.refund_admin_fees, month.delivery_labels, month.adjustments,
            month.residual_expense,
        ]):
            S.put(ws, line, 3 + offset, value, fill=stripe, fmt=S.MONEY,
                  align=S.RIGHT)
        S.put(ws, line, 12, f"=SUM(C{line}:K{line})", font=S.TOTAL_FONT,
              fill=stripe, fmt=S.MONEY, align=S.RIGHT)
        S.put(ws, line, 13, month.other_fees, fill=stripe, fmt=S.MONEY,
              align=S.RIGHT)
        S.put(ws, line, 14, f"=L{line}-M{line}", fill=stripe, fmt=S.MONEY,
              align=S.RIGHT)
    last = detail_first + len(model.months)
    S.put(ws, last, 2, "TOTAL", font=S.TOTAL_FONT, fill=S.TOTAL_FILL,
          align=S.LEFT, border=S.TOP_RULE)
    for col in range(3, 15):
        letter = get_column_letter(col)
        S.put(ws, last, col, f"=SUM({letter}{detail_first}:{letter}{last - 1})",
              font=S.TOTAL_FONT, fill=S.TOTAL_FILL, fmt=S.MONEY,
              align=S.RIGHT, border=S.TOP_RULE)
    S.freeze(ws, 7, 3)


def _write_traffic(wb, model: PLModel):
    ws = _sheet(wb, "Traffic & Conversion")
    months = len(model.months)
    ttm_col = 2 + months
    S.wipe(ws, 1, 30, 1, ttm_col + 2)
    S.widths(ws, {"A": 34})
    S.numeric_widths(ws, 2, ttm_col, 12.0)

    S.title(ws, 1, "TRAFFIC & CONVERSION — from the Business Reports", ttm_col)
    S.subtitle(ws, 2,
               "Conversion = units ordered ÷ sessions. Buy Box is session-weighted "
               "across child ASINs.", ttm_col)
    S.header(ws, 3, ["Metric"] + [m.label for m in model.months] + ["TTM"])

    ttm_start = 2 + (months - len(model.ttm))
    span = f"{get_column_letter(ttm_start)}{{row}}:{get_column_letter(ttm_col - 1)}{{row}}"

    def line(row: int, label: str, values, fmt, summable: bool):
        S.put(ws, row, 1, label, font=S.BOLD, align=S.LEFT)
        for index, value in enumerate(values):
            S.put(ws, row, 2 + index, value, font=S.INPUT, fmt=fmt, align=S.RIGHT)
        if summable:
            S.put(ws, row, ttm_col, f"=SUM({span.format(row=row)})",
                  font=S.TOTAL_FONT, fill=S.TOTAL_FILL, fmt=fmt, align=S.RIGHT)
        else:
            S.put(ws, row, ttm_col, "—", fill=S.TOTAL_FILL, align=S.CENTRE)

    line(4, "Active child ASINs", [m.child_asins for m in model.months], S.INT, False)
    line(5, "Sessions", [m.sessions for m in model.months], S.INT, True)
    line(6, "Page views", [m.page_views for m in model.months], S.INT, True)
    line(7, "Units ordered", [m.units_ordered for m in model.months], S.INT, True)

    S.put(ws, 8, 1, "Conversion % (units ÷ sessions)", font=S.BOLD, align=S.LEFT)
    for index in range(months):
        letter = get_column_letter(2 + index)
        S.put(ws, 8, 2 + index, f"=IFERROR({letter}7/{letter}5,0)",
              fmt=S.PERCENT, align=S.RIGHT)
    S.put(ws, 8, ttm_col,
          f"=IFERROR({get_column_letter(ttm_col)}7/{get_column_letter(ttm_col)}5,0)",
          font=S.TOTAL_FONT, fill=S.TOTAL_FILL, fmt=S.PERCENT, align=S.RIGHT)
    line(9, "Buy Box % (session-weighted)", [m.buy_box for m in model.months],
         S.PERCENT, False)

    traffic = [m for m in model.ttm if m.has_traffic]
    notes: list[str] = []
    if traffic:
        sessions = sum(m.sessions for m in traffic)
        ordered = sum(m.units_ordered for m in traffic)
        if sessions:
            notes.append(f"TTM conversion is {ordered / sessions:.1%} across "
                         f"{sessions:,} sessions.")
        best = max(traffic, key=lambda m: m.units_ordered)
        notes.append(f"Peak month by units ordered: {best.label} "
                     f"({best.units_ordered:,}).")
        low = [m for m in traffic if m.buy_box < 0.95]
        notes.append(
            "Buy Box below 95% in "
            + ", ".join(f"{m.label} ({m.buy_box:.0%})" for m in low)
            + " — investigate a hijacker, pricing suppression or a stock-out."
            if low else "Buy Box held above 95% in every month with a report."
        )
    missing = [m.label for m in model.months if not m.has_traffic]
    if missing:
        notes.append("No Business Report mapped to: " + ", ".join(missing) + ".")

    row = 11
    S.table_title(ws, row, "READ-OUTS", 1, ttm_col)
    for index, text in enumerate(notes, start=1):
        S.note(ws, row + index, text, 1, ttm_col)
    S.freeze(ws, 4, 2)


# ---------------------------------------------------------------------------

SHEET_ORDER = [
    "Status & Missing", "Instructions", "Inputs_Monthly", "Inputs_Units",
    "Inputs_Traffic", "P&L", "SKU Concentration", "COGS Build", "PPC Invoices",
    "Fee Reconciliation", "Traffic & Conversion", "Summary",
]


def build_workbook(model: PLModel, template_path: str, output_path: str) -> str:
    """Fill the template and append the analysis tabs; returns *output_path*."""
    wb = load_workbook(template_path)
    for name in ("Inputs_Monthly", "Inputs_Units", "Inputs_Traffic", "P&L", "Summary"):
        if name in wb.sheetnames:
            S.unmerge(wb[name])

    _fill_inputs_monthly(wb["Inputs_Monthly"], model)
    units_row, cogs_row = _fill_inputs_units(wb["Inputs_Units"], model)
    _fill_inputs_traffic(wb["Inputs_Traffic"], model)
    _fill_pl(wb["P&L"], model, units_row, cogs_row)
    _fill_summary(wb["Summary"], model)

    _write_status(wb, model)
    _write_sku_concentration(wb, model)
    _write_cogs_build(wb, model)
    _write_ppc(wb, model)
    _write_fee_reconciliation(wb, model)
    _write_traffic(wb, model)

    if "Instructions" in wb.sheetnames:
        wb["Instructions"].column_dimensions["B"].width = 120

    wb._sheets = [wb[name] for name in SHEET_ORDER if name in wb.sheetnames] + [
        ws for ws in wb.worksheets if ws.title not in SHEET_ORDER
    ]
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
    wb.active = 0

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    return output_path
